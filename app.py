import os
import json
import re
from io import BytesIO
from uuid import UUID
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, flash, abort, send_file
from dotenv import load_dotenv
from supabase import create_client

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cargar .env junto al archivo (a prueba de PyCharm/OneDrive)
load_dotenv(Path(__file__).with_name(".env"))

SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip()
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()

if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
    raise RuntimeError(
        "Faltan variables de entorno. Define SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY (service role). Revisa .env.example."
    )

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)

def as_int(val, default=None):
    try:
        return int(str(val).strip())
    except Exception:
        return default

def clamp(n, lo, hi):
    return max(lo, min(hi, n))

def fetch_enlace(enlace_id: str):
    try:
        UUID(enlace_id)
    except Exception:
        return None
    resp = supabase.table("enlaces").select("*").eq("id", enlace_id).limit(1).execute()
    rows = resp.data or []
    return rows[0] if rows else None

def fetch_hilos(enlace_id: str):
    resp = supabase.table("hilos").select("*").eq("enlace_id", enlace_id).order("nro_hilo").execute()
    return resp.data or []

def load_programmed_links():
    """
    Devuelve:
    - groups: [{label, items:[{label,value}]}]
    - sites: [sitio1, sitio2, ...]
    """
    p = Path(__file__).with_name("data") / "enlaces_programados.json"
    if not p.exists():
        return {"groups": [], "sites": []}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {"groups": [], "sites": []}

def extract_ruta(nombre_enlace_value: str) -> str:
    """
    Si viene 'ANILLO ESTE | POLO1 - ZARATE', devuelve 'POLO1 - ZARATE'.
    Si no, devuelve el mismo string.
    """
    if not nombre_enlace_value:
        return ""
    s = str(nombre_enlace_value).strip()
    if "|" in s:
        return s.split("|", 1)[1].strip()
    return s

def split_sites(nombre_enlace_value: str):
    ruta = extract_ruta(nombre_enlace_value)
    parts = [p.strip() for p in re.split(r"\s*-\s*", ruta) if p and str(p).strip()]
    if len(parts) == 2:
        return parts[0], parts[1]
    if len(parts) > 2:
        return parts[0], parts[-1]
    return (parts[0], "") if parts else ("", "")

def safe_str(x):
    return (x or "").strip() if isinstance(x, str) else ("" if x is None else str(x))

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-change-me")


@app.get("/")
def dashboard():
    resp = supabase.table("enlaces").select("*").order("creado_at", desc=True).execute()
    enlaces = resp.data or []
    incompletos = [e for e in enlaces if not bool(e.get("completado_b"))]
    validados = [e for e in enlaces if bool(e.get("completado_b"))]
    return render_template("dashboard.html", incompletos=incompletos, validados=validados)


@app.route("/nuevo", methods=["GET", "POST"])
def nuevo_enlace():
    programmed = load_programmed_links()

    if request.method == "GET":
        return render_template(
            "nuevo.html",
            capacidad=24,
            link_groups=programmed.get("groups", []),
            sites=programmed.get("sites", []),
        )

    # nombre de enlace: select o manual
    nombre_sel = safe_str(request.form.get("nombre_enlace_sel"))
    nombre_manual = safe_str(request.form.get("nombre_enlace_manual"))
    nombre_enlace = nombre_manual or nombre_sel

    if not nombre_enlace:
        flash("El Nombre del Enlace es obligatorio.", "danger")
        return redirect(url_for("nuevo_enlace"))

    tipo_cable = safe_str(request.form.get("tipo_cable"))
    capacidad = clamp(as_int(request.form.get("capacidad"), 24) or 24, 1, 288)
    longitud_total = safe_str(request.form.get("longitud_total"))

    # Extremo A (permitimos sugerir desde el enlace si vienen vacíos)
    origen_a = safe_str(request.form.get("origen_a"))
    sala_a = safe_str(request.form.get("sala_a"))
    rack_a = safe_str(request.form.get("rack_a"))
posicion_a = safe_str(request.form.get("posicion_a"))

    # Sugerencia de A/B desde el nombre del enlace (split por "-")
    sA, sB = split_sites(nombre_enlace)
    if not origen_a and sA:
        origen_a = sA

    # Guardamos origen_b "sugerido" desde el enlace, pero seguirá incompleto hasta completar B
    origen_b = safe_str(request.form.get("origen_b_sugerido")) or sB

    # Detalle Extremo A
    hilos_payload = []
    for i in range(1, capacidad + 1):
        desc_a = safe_str(request.form.get(f"desc_a_{i}"))
        hilos_payload.append({"nro_hilo": i, "desc_a": desc_a, "desc_b": None})

    enlace_payload = {
        "nombre_enlace": nombre_enlace,
        "tipo_cable": tipo_cable or None,
        "capacidad": capacidad,
        "longitud_total": longitud_total or None,
        "origen_a": origen_a or None,
        "sala_a": sala_a or None,
        "rack_a": rack_a or None,
        "posicion_a": posicion_a or None,
        "origen_b": origen_b or None,  # sugerido
        "sala_b": None,
        "rack_b": None,
        "completado_b": False,
    }

    try:
        ins = supabase.table("enlaces").insert(enlace_payload).execute()
        row = (ins.data or [None])[0]
        enlace_id = row["id"]
    except Exception as e:
        flash(f"Error creando enlace: {e}", "danger")
        return redirect(url_for("nuevo_enlace"))

    try:
        for r in hilos_payload:
            r["enlace_id"] = enlace_id
        supabase.table("hilos").insert(hilos_payload).execute()
    except Exception as e:
        try:
            supabase.table("enlaces").delete().eq("id", enlace_id).execute()
        except Exception:
            pass
        flash(f"Error creando hilos: {e}", "danger")
        return redirect(url_for("nuevo_enlace"))

    flash("Enlace creado. Quedó en estado Incompleto (pendiente de Extremo B).", "success")
    return redirect(url_for("dashboard"))


@app.route("/enlace/<enlace_id>/b", methods=["GET", "POST"])
def completar_extremo_b(enlace_id):
    programmed = load_programmed_links()
    enlace = fetch_enlace(enlace_id)
    if not enlace:
        abort(404)

    hilos = fetch_hilos(enlace_id)

    if request.method == "GET":
        return render_template("completar_b.html", enlace=enlace, hilos=hilos, sites=programmed.get("sites", []))

    # Datos Extremo B (editable)
    origen_b = safe_str(request.form.get("origen_b"))
    sala_b = safe_str(request.form.get("sala_b"))
    rack_b = safe_str(request.form.get("rack_b"))
    posicion_b = safe_str(request.form.get("posicion_b"))

    try:
        supabase.table("enlaces").update({
            "origen_b": origen_b or None,
            "sala_b": sala_b or None,
            "rack_b": rack_b or None,
            "posicion_b": posicion_b or None,
            "completado_b": True
        }).eq("id", enlace_id).execute()
    except Exception as e:
        flash(f"Error actualizando Extremo B: {e}", "danger")
        return redirect(url_for("completar_extremo_b", enlace_id=enlace_id))

    try:
        updates = []
        for row in hilos:
            nro = row.get("nro_hilo")
            desc_b = safe_str(request.form.get(f"desc_b_{nro}"))
            updates.append({
                "id": row["id"],
                "enlace_id": row["enlace_id"],
                "nro_hilo": nro,
                "desc_a": row.get("desc_a"),
                "desc_b": desc_b
            })
        if updates:
            supabase.table("hilos").upsert(updates).execute()
    except Exception as e:
        flash(f"Extremo B guardado, pero hubo error actualizando hilos: {e}", "warning")

    flash("Enlace Validado correctamente (Extremo B completado).", "success")
    return redirect(url_for("ver_reporte", enlace_id=enlace_id))


@app.route("/enlace/<enlace_id>/editar", methods=["GET", "POST"])
def editar_enlace(enlace_id):
    programmed = load_programmed_links()
    enlace = fetch_enlace(enlace_id)
    if not enlace:
        abort(404)

    hilos = fetch_hilos(enlace_id)

    if request.method == "GET":
        return render_template(
            "editar.html",
            enlace=enlace,
            hilos=hilos,
            link_groups=programmed.get("groups", []),
            sites=programmed.get("sites", []),
        )

    # nombre: select o manual (si manual tiene algo, manda)
    nombre_sel = safe_str(request.form.get("nombre_enlace_sel"))
    nombre_manual = safe_str(request.form.get("nombre_enlace_manual"))
    nombre_enlace = nombre_manual or nombre_sel or enlace.get("nombre_enlace")

    tipo_cable = safe_str(request.form.get("tipo_cable"))
    longitud_total = safe_str(request.form.get("longitud_total"))

    origen_a = safe_str(request.form.get("origen_a"))
    sala_a = safe_str(request.form.get("sala_a"))
    rack_a = safe_str(request.form.get("rack_a"))
posicion_a = safe_str(request.form.get("posicion_a"))

    origen_b = safe_str(request.form.get("origen_b"))
    sala_b = safe_str(request.form.get("sala_b"))
    rack_b = safe_str(request.form.get("rack_b"))

    new_cap = clamp(as_int(request.form.get("capacidad"), enlace.get("capacidad") or 24) or 24, 1, 288)
    old_cap = int(enlace.get("capacidad") or 24)

    # 1) Update cabecera
    try:
        supabase.table("enlaces").update({
            "nombre_enlace": nombre_enlace,
            "tipo_cable": tipo_cable or None,
            "capacidad": new_cap,
            "longitud_total": longitud_total or None,
            "origen_a": origen_a or None,
            "sala_a": sala_a or None,
            "rack_a": rack_a or None,
            "origen_b": origen_b or None,
            "sala_b": sala_b or None,
            "rack_b": rack_b or None,
        }).eq("id", enlace_id).execute()
    except Exception as e:
        flash(f"Error actualizando cabecera: {e}", "danger")
        return redirect(url_for("editar_enlace", enlace_id=enlace_id))

    # 2) Ajustar hilos si cambió capacidad
    try:
        if new_cap > old_cap:
            existing_nums = {int(r.get("nro_hilo")) for r in hilos}
            to_insert = []
            for i in range(old_cap + 1, new_cap + 1):
                if i in existing_nums:
                    continue
                to_insert.append({"enlace_id": enlace_id, "nro_hilo": i, "desc_a": None, "desc_b": None})
            if to_insert:
                supabase.table("hilos").insert(to_insert).execute()

        elif new_cap < old_cap:
            supabase.table("hilos").delete().eq("enlace_id", enlace_id).gt("nro_hilo", new_cap).execute()
    except Exception as e:
        flash(f"Cabecera guardada, pero hubo error ajustando hilos por capacidad: {e}", "warning")

    # 3) Upsert descripciones (A y B) hasta new_cap
    try:
        # recargar hilos (por si se insertaron/borraron)
        hilos = fetch_hilos(enlace_id)
        updates = []
        for row in hilos:
            nro = int(row.get("nro_hilo") or 0)
            if nro < 1 or nro > new_cap:
                continue
            desc_a = safe_str(request.form.get(f"desc_a_{nro}"))
            desc_b = safe_str(request.form.get(f"desc_b_{nro}"))
            updates.append({
                "id": row["id"],
                "enlace_id": enlace_id,
                "nro_hilo": nro,
                "desc_a": desc_a,
                "desc_b": desc_b
            })
        if updates:
            supabase.table("hilos").upsert(updates).execute()
    except Exception as e:
        flash(f"Guardado parcial: error actualizando descripciones: {e}", "warning")

    flash("Cambios guardados correctamente.", "success")
    return redirect(url_for("ver_reporte", enlace_id=enlace_id))


@app.get("/enlace/<enlace_id>/ver")
def ver_reporte(enlace_id):
    enlace = fetch_enlace(enlace_id)
    if not enlace:
        abort(404)
    hilos = fetch_hilos(enlace_id)

    diffs = 0
    for h in hilos:
        a = safe_str(h.get("desc_a"))
        b = safe_str(h.get("desc_b"))
        if a != b:
            diffs += 1

    # nombres cortos para la vista tipo excel
    a_site = enlace.get("origen_a") or split_sites(enlace.get("nombre_enlace") or "")[0] or "Extremo A"
    b_site = enlace.get("origen_b") or split_sites(enlace.get("nombre_enlace") or "")[1] or "Extremo B"

    return render_template("ver.html", enlace=enlace, hilos=hilos, diffs=diffs, a_site=a_site, b_site=b_site)


def build_xlsx(enlace: dict, hilos: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ENLACE"

    thin = Side(style="thin")
    thick = Side(style="medium")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    # Column widths
    widths = {"A": 8, "B": 55, "C": 55, "D": 8, "E": 3, "F": 26, "G": 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    a_site = enlace.get("origen_a") or split_sites(enlace.get("nombre_enlace") or "")[0] or "EXTREMO A"
    b_site = enlace.get("origen_b") or split_sites(enlace.get("nombre_enlace") or "")[1] or "EXTREMO B"

    # Headers like template
    ws.merge_cells("A2:B2")
    ws["A2"] = str(a_site).upper()
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C2:D2")
    ws["C2"] = str(b_site).upper()
    ws["C2"].font = Font(bold=True)
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "Fibra"
    ws["B3"] = "DESCRIPCION"
    ws["C3"] = "DESCRIPCION"
    ws["D3"] = "Fibra"
    for cell in ("A3", "B3", "C3", "D3"):
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    start_row = 4
    cap = int(enlace.get("capacidad") or len(hilos) or 24)

    # Ensure list ordered
    hilos_sorted = sorted(hilos, key=lambda x: int(x.get("nro_hilo") or 0))
    for i in range(1, cap + 1):
        r = start_row + (i - 1)
        # find hilo row
        row = next((h for h in hilos_sorted if int(h.get("nro_hilo") or 0) == i), None)
        desc_a = safe_str(row.get("desc_a")) if row else ""
        desc_b = safe_str(row.get("desc_b")) if row else ""

        ws[f"A{r}"] = i
        ws[f"B{r}"] = desc_a
        ws[f"C{r}"] = desc_b
        ws[f"D{r}"] = i

        ws[f"A{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")

        # Borders thin
        for c in ("A", "B", "C", "D"):
            ws[f"{c}{r}"].border = border_thin

    # borders for headers
    for c in ("A", "B", "C", "D"):
        ws[f"{c}3"].border = border_thin
    ws["A2"].border = border_thin
    ws["C2"].border = border_thin

    # Thick outline around main table
    top = 2
    left = 1
    right = 4
    bottom = start_row + cap - 1
    for col in range(left, right + 1):
        ws.cell(row=top, column=col).border = Border(
            left=thick if col == left else thin,
            right=thick if col == right else thin,
            top=thick,
            bottom=thin
        )
    for row in range(top + 1, bottom):
        ws.cell(row=row, column=left).border = Border(left=thick, right=thin, top=thin, bottom=thin)
        ws.cell(row=row, column=right).border = Border(left=thin, right=thick, top=thin, bottom=thin)
    for col in range(left, right + 1):
        ws.cell(row=bottom, column=col).border = Border(
            left=thick if col == left else thin,
            right=thick if col == right else thin,
            top=thin,
            bottom=thick
        )

    # Summary box helper
    def write_box(top_row: int, title: str, origin: str, sala: str, cable: str, rack: str, posicion: str):
        labels = [
            ("ENLACE:", safe_str(enlace.get("nombre_enlace"))),
            ("ORIGEN", origin),
            ("SALA", sala),
            ("CABLE", cable),
            ("RACK", rack),
            ("POSICIÓN", posicion),
            ("CAPACIDAD", safe_str(enlace.get("capacidad"))),
            ("LONGITUD TRAMO TOTAL", safe_str(enlace.get("longitud_total"))),
        ]
        r = top_row
        for k, v in labels:
            ws[f"F{r}"] = k
            ws[f"G{r}"] = v
            ws[f"F{r}"].font = Font(bold=True)
            ws[f"F{r}"].alignment = Alignment(vertical="center")
            ws[f"G{r}"].alignment = Alignment(vertical="center", wrap_text=True)
            ws[f"F{r}"].border = border_thin
            ws[f"G{r}"].border = border_thin
            r += 1

        # Thick outline
        box_bottom = top_row + len(labels) - 1
        for rr in range(top_row, box_bottom + 1):
            ws[f"F{rr}"].border = Border(left=thick, right=thin, top=thick if rr == top_row else thin, bottom=thick if rr == box_bottom else thin)
            ws[f"G{rr}"].border = Border(left=thin, right=thick, top=thick if rr == top_row else thin, bottom=thick if rr == box_bottom else thin)

    write_box(
        top_row=4,
        title="A",
        origin=safe_str(enlace.get("origen_a") or a_site),
        sala=safe_str(enlace.get("sala_a")),
        cable=safe_str(enlace.get("tipo_cable")),
        rack=safe_str(enlace.get("rack_a")),
        posicion=safe_str(enlace.get("posicion_a")),
    )
    write_box(
        top_row=13,
        title="B",
        origin=safe_str(enlace.get("origen_b") or b_site),
        sala=safe_str(enlace.get("sala_b")),
        cable=safe_str(enlace.get("tipo_cable")),
        rack=safe_str(enlace.get("rack_b")),
        posicion=safe_str(enlace.get("posicion_b")),
    )

    # Save bytes
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.get("/enlace/<enlace_id>/export.xlsx")
def export_xlsx(enlace_id):
    enlace = fetch_enlace(enlace_id)
    if not enlace:
        abort(404)
    hilos = fetch_hilos(enlace_id)

    content = build_xlsx(enlace, hilos)

    # filename safe
    name = safe_str(enlace.get("nombre_enlace")) or "enlace"
    name = name.replace("/", "-").replace("\\", "-").replace(":", "-")
    filename = f"{name}.xlsx"

    return send_file(
        BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.getenv("PORT", "5000")))
